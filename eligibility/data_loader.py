"""
Data Loader - Load and validate eligibility data files.

Loads 2 Excel data files at startup:
- eligible_customers.xlsx (eligible accounts lookup)
- reasons_file.xlsx (ineligible accounts + evidence)

Creates indexes on account_number for O(1) lookups.
Caches all data in memory for fast access.
Raises exceptions if any file is missing or malformed.
"""

import os
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError(
        "openpyxl is required. Install with: pip install openpyxl"
    )

from logger.rag_logging import RAGLogger
from logger.session_manager import SessionManager


class DataLoader:
    """Singleton data loader with validation and caching."""

    _instance: Optional["DataLoader"] = None

    def __new__(cls) -> "DataLoader":
        """Ensure singleton pattern."""
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        """Initialize data loader and load all data files."""
        if self._initialized:
            return

        self._initialized = True
        self.rag_logger = RAGLogger()
        self.session_manager = SessionManager()

        # Get data directory path
        self.data_dir = os.path.join(
            os.path.dirname(__file__), "data"
        )

        # Cache for all data
        self.eligible_customers: Dict[str, Dict[str, Any]] = {}
        self.reasons_file: Dict[str, Dict[str, Any]] = {}

        # Load all data at startup
        self._load_all_data()

    def _load_all_data(self) -> None:
        """Load and validate all data files."""
        start_time = datetime.now(timezone.utc)
        request_id = self.rag_logger.generate_request_id()

        try:
            # Load eligible_customers.xlsx
            self.eligible_customers = self._load_excel_file(
                "eligible_customers.xlsx",
                request_id,
                "ACCOUNTNO"
            )

            # Load reasons_file.xlsx
            self.reasons_file = self._load_excel_file(
                "reasons_file.xlsx",
                request_id,
                "account_number"
            )

            latency_ms = (
                datetime.now(timezone.utc) - start_time
            ).total_seconds() * 1000

            # Log successful load
            self.rag_logger.log(
                request_id=request_id,
                event="eligibility_data_load_success",
                severity="INFO",
                message="Eligibility data files loaded successfully",
                context={
                    "eligible_customers_rows": len(self.eligible_customers),
                    "reasons_file_rows": len(self.reasons_file),
                    "latency_ms": latency_ms,
                    "data_dir": self.data_dir,
                }
            )

        except Exception as e:
            # Log critical failure
            self.rag_logger.log(
                request_id=request_id,
                event="eligibility_data_load_failure",
                severity="CRITICAL",
                message=f"Failed to load eligibility data files: {str(e)}",
                context={
                    "error_type": type(e).__name__,
                    "error_message": str(e),
                    "data_dir": self.data_dir,
                }
            )
            raise

    def _load_excel_file(
        self,
        filename: str,
        request_id: str,
        key_column: str
    ) -> Dict[str, Dict[str, Any]]:
        """
        Load and validate an Excel data file.

        Args:
            filename: Name of the Excel file to load.
            request_id: Request ID for logging.
            key_column: Column name to use as lookup key.

        Returns:
            Dictionary indexed by key_column value.

        Raises:
            FileNotFoundError: If file does not exist.
            Exception: If file is malformed or key_column not found.
        """
        filepath = os.path.join(self.data_dir, filename)

        if not os.path.exists(filepath):
            self.rag_logger.log(
                request_id=request_id,
                event="data_file_not_found",
                severity="ERROR",
                message=f"Data file not found: {filepath}",
                context={"filename": filename, "filepath": filepath}
            )
            raise FileNotFoundError(f"Data file not found: {filepath}")

        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            worksheet = workbook.active

            if worksheet is None:
                raise ValueError(f"No active worksheet found in {filepath}")

            # Get header row (assume first row)
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)

            # Find key column index
            try:
                key_col_idx = headers.index(key_column)
            except ValueError:
                self.rag_logger.log(
                    request_id=request_id,
                    event="key_column_not_found",
                    severity="ERROR",
                    message=f"Key column '{key_column}' not found in {filename}",
                    context={
                        "filename": filename,
                        "key_column": key_column,
                        "available_columns": headers,
                    }
                )
                raise

            # Load data rows
            data_dict: Dict[str, Dict[str, Any]] = {}
            duplicate_keys = []

            for row_idx, row in enumerate(worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                values_only=True
            ), start=2):
                if not row or all(cell is None for cell in row):
                    continue  # Skip empty rows

                # Build row dict
                row_dict = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < len(row):
                        row_dict[header] = row[col_idx]

                # Extract key
                key_value = str(row[key_col_idx]) if key_col_idx < len(row) else None
                if not key_value or key_value == "None":
                    continue  # Skip rows without key

                # Check for duplicates
                if key_value in data_dict:
                    duplicate_keys.append(key_value)
                else:
                    data_dict[key_value] = row_dict

            workbook.close()

            # Log duplicate warnings if any
            if duplicate_keys:
                self.rag_logger.log(
                    request_id=request_id,
                    event="data_file_duplicates_found",
                    severity="WARNING",
                    message=f"Duplicate key values found in {filename}",
                    context={
                        "filename": filename,
                        "duplicate_count": len(duplicate_keys),
                        "key_column": key_column,
                    }
                )

            self.rag_logger.log(
                request_id=request_id,
                event="data_file_loaded",
                severity="DEBUG",
                message=f"Data file loaded: {filename}",
                context={
                    "filename": filename,
                    "rows_loaded": len(data_dict),
                    "duplicates_found": len(duplicate_keys),
                }
            )

            return data_dict

        except Exception as e:
            self.rag_logger.log(
                request_id=request_id,
                event="data_file_load_error",
                severity="ERROR",
                message=f"Error loading data file {filename}: {str(e)}",
                context={
                    "filename": filename,
                    "error_type": type(e).__name__,
                    "error_message": str(e),
                }
            )
            raise

    def get_eligible_customer(self, account_number: str) -> Optional[Dict[str, Any]]:
        """
        Get eligible customer record by account number.

        Args:
            account_number: 10-digit account number.

        Returns:
            Customer record dict if found, None otherwise.
        """
        return self.eligible_customers.get(account_number)

    def get_reasons_record(self, account_number: str) -> Optional[Dict[str, Any]]:
        """
        Get reasons record by account number.

        Args:
            account_number: 10-digit account number.

        Returns:
            Reasons record dict if found, None otherwise.
        """
        return self.reasons_file.get(account_number)

    def is_eligible(self, account_number: str) -> bool:
        """Check if account is in eligible customers list."""
        return account_number in self.eligible_customers

    def has_ineligibility_reasons(self, account_number: str) -> bool:
        """Check if account is in reasons file (ineligible)."""
        return account_number in self.reasons_file

    def get_all_eligible_accounts(self) -> List[str]:
        """Get list of all eligible account numbers."""
        return list(self.eligible_customers.keys())

    def get_all_ineligible_accounts(self) -> List[str]:
        """Get list of all ineligible account numbers."""
        return list(self.reasons_file.keys())

    def get_data_summary(self) -> Dict[str, Any]:
        """Get summary of loaded data."""
        return {
            "eligible_customers_count": len(self.eligible_customers),
            "reasons_file_count": len(self.reasons_file),
            "total_accounts": len(self.eligible_customers) + len(self.reasons_file),
        }
